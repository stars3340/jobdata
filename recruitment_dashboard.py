import dash
from dash import dcc, html, Input, Output, State, callback_context, dash_table
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import pymysql
from datetime import datetime, timedelta, date
import warnings
import time
import io

# 过滤警告
warnings.filterwarnings('ignore')

# 数据库连接配置 - 使用统一配置管理
import os
from config import Config

DB_CONFIG = Config.DB_CONFIG

def get_db_connection():
    """获取数据库连接"""
    try:
        # 添加连接参数优化
        config = DB_CONFIG.copy()
        config.update({
            'connect_timeout': 10,
            'read_timeout': 30,
            'write_timeout': 30,
            'autocommit': True
        })
        return pymysql.connect(**config)
    except Exception as e:
        print(f"数据库连接失败: {e}")
        return None

def query_data(sql):
    """执行SQL查询并返回DataFrame"""
    connection = get_db_connection()
    if connection:
        try:
            df = pd.read_sql(sql, connection)
            return df
        except Exception as e:
            print(f"查询执行失败: {e}")
            return pd.DataFrame()
        finally:
            connection.close()
    return pd.DataFrame()

# 创建Dash应用
app = dash.Dash(__name__, external_stylesheets=['assets/style.css'])
app.title = "智能招聘数据分析平台"

# 定义颜色主题
colors = {
    'primary': '#06D6A0',
    'secondary': '#118AB2',
    'success': '#06D6A0',
    'warning': '#FFD166',
    'danger': '#EF476F',
    'background': '#0A0E1A',
    'text': '#FFFFFF'
}

def create_loading_component():
    """创建加载组件"""
    return html.Div([
        html.Div(className="loading-spinner"),
        html.Span("数据加载中...", style={'marginLeft': '10px', 'color': colors['text']})
    ], style={'display': 'flex', 'alignItems': 'center', 'justifyContent': 'center', 'padding': '20px'})

def get_funnel_data(start_date=None, end_date=None, user_id=None):
    """获取漏斗数据"""
    where_conditions = []
    
    if start_date and end_date:
        where_conditions.append(f"create_time BETWEEN '{start_date}' AND '{end_date} 23:59:59'")
    
    if user_id and user_id != 'all':
        where_conditions.append(f"uid = '{user_id}'")
    
    where_clause = " AND ".join(where_conditions)
    if where_clause:
        where_clause = f"WHERE {where_clause}"
    
    sql = f"""
    SELECT 
        event_type,
        COUNT(*) as count
    FROM recruit_event 
    {where_clause}
    GROUP BY event_type
    ORDER BY count DESC
    """
    
    df = query_data(sql)
    
    # 事件类型映射（数字到中文）- 根据Prisma schema修正
    event_type_mapping = {
        '1': '浏览简历',        # VIEW_RESUME
        '2': '打招呼',         # SCREEN_PASS  
        '12': '相互沟通',       # BOSS_CHAT
        '13': '建联量'          # PHONE_NUMBER
    }
    
    # 重新定义漏斗顺序
    funnel_mapping = {
        '浏览简历': ('浏览简历', 1),
        '打招呼': ('打招呼', 2), 
        '相互沟通': ('相互沟通', 3),
        '建联量': ('建联量', 4)
    }
    
    funnel_data = []
    for event_num, event_name in event_type_mapping.items():
        count = df[df['event_type'] == event_num]['count'].sum() if not df.empty else 0
        if event_name in funnel_mapping:
            stage_name, order = funnel_mapping[event_name]
            funnel_data.append({
                'stage': stage_name,
                'count': count,
                'order': order
            })
    
    return pd.DataFrame(funnel_data).sort_values('order')

def get_trend_data(start_date=None, end_date=None, user_id=None):
    """获取趋势数据"""
    where_conditions = []
    
    if start_date and end_date:
        where_conditions.append(f"create_time BETWEEN '{start_date}' AND '{end_date} 23:59:59'")
    
    if user_id and user_id != 'all':
        where_conditions.append(f"uid = '{user_id}'")
    
    where_clause = " AND ".join(where_conditions)
    if where_clause:
        where_clause = f"WHERE {where_clause}"
    
    sql = f"""
    SELECT 
        DATE(create_time) as date,
        event_type,
        COUNT(*) as count
    FROM recruit_event 
    {where_clause}
    GROUP BY DATE(create_time), event_type
    ORDER BY date DESC
    LIMIT 100
    """
    
    df = query_data(sql)
    
    # 添加事件类型映射 - 根据Prisma schema修正
    if not df.empty:
        event_type_mapping = {
            '1': '浏览简历',        # VIEW_RESUME
            '2': '打招呼',         # SCREEN_PASS
            '12': '相互沟通',       # BOSS_CHAT
            '13': '建联量'          # PHONE_NUMBER
        }
        df['event_type'] = df['event_type'].map(event_type_mapping).fillna(df['event_type'])
    
    return df

def get_user_list():
    """获取用户列表"""
    sql = """
    SELECT 
        u.id,
        CASE 
            WHEN u.name IS NOT NULL AND u.name != '' THEN u.name
            ELSE CONCAT('用户-', LEFT(u.id, 8))
        END as display_name,
        COUNT(re.id) as event_count
    FROM user u
    LEFT JOIN recruit_event re ON u.id = re.uid
    GROUP BY u.id, u.name
    HAVING event_count > 0
    ORDER BY event_count DESC
    """
    df = query_data(sql)
    
    options = [{'label': '全部用户', 'value': 'all'}]
    
    if not df.empty:
        for _, row in df.iterrows():
            # 显示用户名和事件数量
            options.append({
                'label': f"{row['display_name']} ({row['event_count']}条)",
                'value': str(row['id'])
            })
    
    return options

def get_key_metrics(start_date=None, end_date=None, user_id=None):
    """获取关键指标"""
    funnel_df = get_funnel_data(start_date, end_date, user_id)
    
    if funnel_df.empty:
        return {
            'browse_resumes': 0,
            'greetings': 0, 
            'mutual_communications': 0,
            'connections': 0,
            'greeting_success_rate': 0,
            'communication_success_rate': 0,
            'mutual_communication_rate': 0,
            'resume_screening_rate': 0
        }
    
    # 获取各阶段数据
    views = funnel_df[funnel_df['stage'] == '浏览简历']['count'].iloc[0] if len(funnel_df[funnel_df['stage'] == '浏览简历']) > 0 else 0
    screening = funnel_df[funnel_df['stage'] == '打招呼']['count'].iloc[0] if len(funnel_df[funnel_df['stage'] == '打招呼']) > 0 else 0
    chats = funnel_df[funnel_df['stage'] == '相互沟通']['count'].iloc[0] if len(funnel_df[funnel_df['stage'] == '相互沟通']) > 0 else 0
    contacts = funnel_df[funnel_df['stage'] == '建联量']['count'].iloc[0] if len(funnel_df[funnel_df['stage'] == '建联量']) > 0 else 0
    
    # 计算转化率
    greeting_success_rate = (contacts / screening * 100) if screening > 0 else 0  # 建联量/打招呼数
    communication_success_rate = (contacts / chats * 100) if chats > 0 else 0  # 建联量/相互沟通数
    mutual_communication_rate = (chats / screening * 100) if screening > 0 else 0  # 相互沟通数/打招呼数
    resume_screening_rate = (screening / views * 100) if views > 0 else 0  # 打招呼数/浏览简历
    
    return {
        'browse_resumes': views,          # 浏览简历
        'greetings': screening,           # 打招呼
        'mutual_communications': chats,   # 相互沟通
        'connections': contacts,          # 建联量
        'greeting_success_rate': greeting_success_rate,           # 打招呼成功率
        'communication_success_rate': communication_success_rate, # 沟通成功率
        'mutual_communication_rate': mutual_communication_rate,   # 相互沟通率
        'resume_screening_rate': resume_screening_rate            # 简历过筛率
    }

def create_metric_card(title, value, change=None, format_type='number', icon=None, calculation_formula=None):
    """创建指标卡片，包含计算方式说明"""
    # 根据格式类型处理数值显示
    if format_type == 'percentage':
        display_value = f"{value:.1f}%"
    elif format_type == 'number':
        display_value = f"{value:,}"
    else:
        display_value = str(value)
    
    # 变化指示器
    change_element = html.Div()
    if change is not None:
        change_class = "positive" if change >= 0 else "negative"
        change_text = f"+{change:.1f}%" if change >= 0 else f"{change:.1f}%"
        change_element = html.Div(
            change_text,
            className=f"kpi-change {change_class}"
        )
    
    # 计算方式说明
    formula_element = html.Div()
    if calculation_formula:
        formula_element = html.Div(
            calculation_formula,
            className="kpi-formula",
            style={
                'fontSize': '0.75rem',
                'color': '#CBD5E1',
                'marginTop': '0.5rem',
                'padding': '0.5rem',
                'backgroundColor': 'rgba(255,255,255,0.05)',
                'borderRadius': '4px',
                'border': '1px solid rgba(255,255,255,0.1)'
            }
        )
    
    return html.Div([
        html.H3(title),
        html.Div(display_value, className="kpi-value"),
        change_element,
        formula_element
    ], className="card kpi-card")

def create_funnel_chart(funnel_df):
    """创建漏斗图"""
    if funnel_df.empty:
        return go.Figure().add_annotation(
            text="暂无数据",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color=colors['text'])
        )
    
    # 计算转化率
    funnel_df = funnel_df.copy()
    funnel_df['conversion_rate'] = 0
    
    for i in range(1, len(funnel_df)):
        if funnel_df.iloc[i-1]['count'] > 0:
            rate = (funnel_df.iloc[i]['count'] / funnel_df.iloc[i-1]['count']) * 100
            funnel_df.iloc[i, funnel_df.columns.get_loc('conversion_rate')] = rate
    
    # 计算各阶段的文字标签
    stage_texts = []
    total_count = funnel_df.iloc[0]['count'] if not funnel_df.empty else 0
    
    for i, (_, row) in enumerate(funnel_df.iterrows()):
        count = row['count']
        if i == 0:
            # 第一阶段
            stage_texts.append(f"{count}")
        else:
            # 后续阶段计算转化率
            prev_count = funnel_df.iloc[i-1]['count']
            if total_count > 0:
                total_percent = count / total_count * 100
                prev_percent = count / prev_count * 100 if prev_count > 0 else 0
                stage_texts.append(f"{count}<br>总体: {total_percent:.1f}%<br>转化: {prev_percent:.1f}%")
            else:
                stage_texts.append(f"{count}")
    
    # 创建漏斗图
    fig = go.Figure(go.Funnel(
        y=funnel_df['stage'],
        x=funnel_df['count'],
        textinfo="text",
        text=stage_texts,
        textfont=dict(size=14, color='white'),
        marker=dict(
            color=[colors['primary'], colors['secondary'], colors['warning'], colors['success']],
            line=dict(width=2, color='white')
        ),
        connector=dict(line=dict(color=colors['primary'], dash='dot', width=3))
    ))
    
    fig.update_layout(
        title=dict(
            text="<b>招聘漏斗分析</b>",
            font=dict(size=20, color=colors['text']),
            x=0.5
        ),
        font=dict(color=colors['text']),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(t=80, b=40, l=40, r=40)
    )
    
    return fig

def create_trend_chart(trend_df):
    """创建趋势图"""
    if trend_df.empty:
        fig = go.Figure()
        fig.add_annotation(
            text="暂无数据",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color=colors['text'])
        )
    else:
        # 透视数据
        pivot_df = trend_df.pivot_table(
            index='date', 
            columns='event_type', 
            values='count', 
            fill_value=0
        ).reset_index()
        
        fig = go.Figure()
        
        # 定义颜色映射
        color_map = {
            '浏览简历': colors['primary'],
            '打招呼': colors['secondary'], 
            '相互沟通': colors['warning'],
            '建联量': colors['success']
        }
        
        # 添加每个事件类型的线条
        for col in pivot_df.columns[1:]:
            if col in color_map:
                fig.add_trace(go.Scatter(
                    x=pivot_df['date'],
                    y=pivot_df[col],
                    mode='lines+markers',
                    name=col,
                    line=dict(color=color_map[col], width=3),
                    marker=dict(size=8, line=dict(width=2, color='white')),
                    hovertemplate=f'<b>{col}</b><br>日期: %{{x}}<br>数量: %{{y}}<extra></extra>'
                ))
    
    fig.update_layout(
        title=dict(
            text="<b>每日活动趋势</b>",
            font=dict(size=20, color=colors['text']),
            x=0.5
        ),
        xaxis=dict(
            title=dict(text="日期", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)'
        ),
        yaxis=dict(
            title=dict(text="活动数量", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)'
        ),
        legend=dict(
            font=dict(color=colors['text']),
            bgcolor='rgba(35, 41, 70, 0.8)',
            bordercolor=colors['primary'],
            borderwidth=1
        ),
        font=dict(color=colors['text']),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(t=80, b=60, l=60, r=40),
        hovermode='x unified'
    )
    
    return fig

def get_detailed_data(start_date=None, end_date=None, user_id=None, limit=None):
    """获取详细数据"""
    where_conditions = []
    
    if start_date and end_date:
        where_conditions.append(f"re.create_time BETWEEN '{start_date}' AND '{end_date} 23:59:59'")
    
    if user_id and user_id != 'all':
        where_conditions.append(f"re.uid = '{user_id}'")
    
    where_clause = " AND ".join(where_conditions)
    if where_clause:
        where_clause = f"WHERE {where_clause}"
    
    # 添加LIMIT控制，导出时不限制，显示时限制1000条
    limit_clause = f"LIMIT {limit}" if limit else ""
    
    sql = f"""
    SELECT 
        re.id,
        re.event_type as '事件类型',
        CASE 
            WHEN u.name IS NOT NULL AND u.name != '' THEN u.name
            WHEN re.uid IS NOT NULL THEN CONCAT('用户-', LEFT(re.uid, 8))
            ELSE '未知用户'
        END as '用户',
        re.resume_id as '简历ID',
        re.job_id as '职位ID', 
        DATE_FORMAT(re.create_time, '%Y-%m-%d %H:%i:%s') as '创建时间'
    FROM recruit_event re
    LEFT JOIN user u ON re.uid = u.id
    {where_clause}
    ORDER BY re.create_time DESC
    {limit_clause}
    """
    
    df = query_data(sql)
    
    # 添加事件类型映射 - 根据Prisma schema修正
    if not df.empty:
        event_type_mapping = {
            '1': '浏览简历',        # VIEW_RESUME
            '2': '打招呼',         # SCREEN_PASS
            '12': '相互沟通',       # BOSS_CHAT
            '13': '建联量'          # PHONE_NUMBER
        }
        df['事件类型'] = df['事件类型'].map(event_type_mapping).fillna(df['事件类型'])
    
    return df

def get_greeting_success_trend(start_date=None, end_date=None, user_id=None):
    """获取打招呼成功率趋势数据"""
    where_conditions = []
    
    if start_date and end_date:
        where_conditions.append(f"create_time BETWEEN '{start_date}' AND '{end_date} 23:59:59'")
    
    if user_id and user_id != 'all':
        where_conditions.append(f"uid = '{user_id}'")
    
    where_clause = " AND ".join(where_conditions)
    if where_clause:
        where_clause = f"WHERE {where_clause}"
    
    sql = f"""
    SELECT 
        DATE(create_time) as date,
        event_type,
        COUNT(*) as count
    FROM recruit_event 
    {where_clause}
    GROUP BY DATE(create_time), event_type
    ORDER BY date ASC
    """
    
    df = query_data(sql)
    
    if df.empty:
        return pd.DataFrame()
    
    # 透视数据
    pivot_df = df.pivot_table(
        index='date', 
        columns='event_type', 
        values='count', 
        fill_value=0
    ).reset_index()
    
    # 计算每日打招呼成功率
    success_rate_data = []
    for _, row in pivot_df.iterrows():
        greetings = row.get('2', 0)  # 打招呼
        connections = row.get('13', 0)  # 建联量
        
        success_rate = (connections / greetings * 100) if greetings > 0 else 0
        success_rate_data.append({
            'date': row['date'],
            'greeting_success_rate': success_rate,
            'greetings': greetings,
            'connections': connections
        })
    
    return pd.DataFrame(success_rate_data)

def get_communication_success_trend(start_date=None, end_date=None, user_id=None):
    """获取沟通成功率趋势数据"""
    where_conditions = []
    
    if start_date and end_date:
        where_conditions.append(f"create_time BETWEEN '{start_date}' AND '{end_date} 23:59:59'")
    
    if user_id and user_id != 'all':
        where_conditions.append(f"uid = '{user_id}'")
    
    where_clause = " AND ".join(where_conditions)
    if where_clause:
        where_clause = f"WHERE {where_clause}"
    
    sql = f"""
    SELECT 
        DATE(create_time) as date,
        event_type,
        COUNT(*) as count
    FROM recruit_event 
    {where_clause}
    GROUP BY DATE(create_time), event_type
    ORDER BY date ASC
    """
    
    df = query_data(sql)
    
    if df.empty:
        return pd.DataFrame()
    
    # 透视数据
    pivot_df = df.pivot_table(
        index='date', 
        columns='event_type', 
        values='count', 
        fill_value=0
    ).reset_index()
    
    # 计算每日沟通成功率
    success_rate_data = []
    for _, row in pivot_df.iterrows():
        mutual_communications = row.get('12', 0)  # 相互沟通
        connections = row.get('13', 0)  # 建联量
        
        success_rate = (connections / mutual_communications * 100) if mutual_communications > 0 else 0
        success_rate_data.append({
            'date': row['date'],
            'communication_success_rate': success_rate,
            'mutual_communications': mutual_communications,
            'connections': connections
        })
    
    return pd.DataFrame(success_rate_data)

def get_mutual_communication_trend(start_date=None, end_date=None, user_id=None):
    """获取相互沟通率趋势数据"""
    where_conditions = []
    
    if start_date and end_date:
        where_conditions.append(f"create_time BETWEEN '{start_date}' AND '{end_date} 23:59:59'")
    
    if user_id and user_id != 'all':
        where_conditions.append(f"uid = '{user_id}'")
    
    where_clause = " AND ".join(where_conditions)
    if where_clause:
        where_clause = f"WHERE {where_clause}"
    
    sql = f"""
    SELECT 
        DATE(create_time) as date,
        event_type,
        COUNT(*) as count
    FROM recruit_event 
    {where_clause}
    GROUP BY DATE(create_time), event_type
    ORDER BY date ASC
    """
    
    df = query_data(sql)
    
    if df.empty:
        return pd.DataFrame()
    
    # 透视数据
    pivot_df = df.pivot_table(
        index='date', 
        columns='event_type', 
        values='count', 
        fill_value=0
    ).reset_index()
    
    # 计算每日相互沟通率
    communication_rate_data = []
    for _, row in pivot_df.iterrows():
        greetings = row.get('2', 0)  # 打招呼
        mutual_communications = row.get('12', 0)  # 相互沟通
        
        communication_rate = (mutual_communications / greetings * 100) if greetings > 0 else 0
        communication_rate_data.append({
            'date': row['date'],
            'mutual_communication_rate': communication_rate,
            'greetings': greetings,
            'mutual_communications': mutual_communications
        })
    
    return pd.DataFrame(communication_rate_data)

def create_greeting_success_trend_chart(trend_df):
    """创建打招呼成功率趋势图"""
    # 创建图表，强制设置Y轴范围
    fig = go.Figure()
    
    # 添加一个隐藏的空数据点来强制Y轴范围
    fig.add_trace(go.Scatter(
        x=[],
        y=[0, 5],  # 强制Y轴从0到5
        mode='markers',
        marker=dict(opacity=0),  # 完全透明
        showlegend=False,
        hoverinfo='skip'
    ))
    
    if trend_df.empty:
        # 暂无数据时显示提示
        fig.add_annotation(
            text="暂无数据",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color=colors['text'])
        )
    else:
        # 添加成功率线条
        fig.add_trace(go.Scatter(
            x=trend_df['date'],
            y=trend_df['greeting_success_rate'],
            mode='lines+markers',
            name='打招呼成功率',
            line=dict(color=colors['success'], width=3),
            marker=dict(size=8, line=dict(width=2, color='white')),
            hovertemplate='<b>打招呼成功率</b><br>日期: %{x}<br>成功率: %{y:.1f}%<br>打招呼: %{customdata[0]}<br>建联: %{customdata[1]}<extra></extra>',
            customdata=list(zip(trend_df['greetings'], trend_df['connections']))
        ))
    
    # 强制布局设置
    fig.update_layout(
        title=dict(
            text="<b>打招呼成功率趋势分析</b>",
            font=dict(size=20, color=colors['text']),
            x=0.5
        ),
        xaxis=dict(
            title=dict(text="日期", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)'
        ),
        yaxis=dict(
            title=dict(text="成功率 (%)", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)',
            range=[0, 5],
            fixedrange=True,
            autorange=False,
            tickformat='.1f',
            dtick=1,  # 每1%一个刻度
            tick0=0,  # 从0开始
            constrain='domain'
        ),
        legend=dict(
            font=dict(color=colors['text']),
            bgcolor='rgba(35, 41, 70, 0.8)',
            bordercolor=colors['primary'],
            borderwidth=1
        ),
        font=dict(color=colors['text']),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(t=80, b=60, l=60, r=40),
        hovermode='x unified',
        autosize=True
    )
    
    # 多重强制设置Y轴
    fig.update_yaxes(
        range=[0, 5],
        autorange=False, 
        fixedrange=True,
        constrain='domain',
        dtick=1,
        tick0=0
    )
    
    # 最后一道保险：直接修改layout
    fig.layout.yaxis.range = [0, 5]
    fig.layout.yaxis.autorange = False
    
    return fig

def create_communication_success_trend_chart(trend_df):
    """创建沟通成功率趋势图"""
    # 创建图表，强制设置Y轴范围
    fig = go.Figure()
    
    # 添加一个隐藏的空数据点来强制Y轴范围
    fig.add_trace(go.Scatter(
        x=[],
        y=[0, 100],  # 强制Y轴从0到100
        mode='markers',
        marker=dict(opacity=0),  # 完全透明
        showlegend=False,
        hoverinfo='skip'
    ))
    
    if trend_df.empty:
        # 暂无数据时显示提示
        fig.add_annotation(
            text="暂无数据",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color=colors['text'])
        )
    else:
        # 添加沟通成功率线条
        fig.add_trace(go.Scatter(
            x=trend_df['date'],
            y=trend_df['communication_success_rate'],
            mode='lines+markers',
            name='沟通成功率',
            line=dict(color=colors['secondary'], width=3),
            marker=dict(size=8, line=dict(width=2, color='white')),
            hovertemplate='<b>沟通成功率</b><br>日期: %{x}<br>成功率: %{y:.1f}%<br>相互沟通: %{customdata[0]}<br>建联: %{customdata[1]}<extra></extra>',
            customdata=list(zip(trend_df['mutual_communications'], trend_df['connections']))
        ))
    
    # 强制布局设置
    fig.update_layout(
        title=dict(
            text="<b>沟通成功率趋势分析</b>",
            font=dict(size=20, color=colors['text']),
            x=0.5
        ),
        xaxis=dict(
            title=dict(text="日期", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)'
        ),
        yaxis=dict(
            title=dict(text="成功率 (%)", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)',
            range=[0, 100],
            fixedrange=True,
            autorange=False,
            tickformat='.1f',
            dtick=20,  # 每20%一个刻度
            tick0=0,  # 从0开始
            constrain='domain'
        ),
        legend=dict(
            font=dict(color=colors['text']),
            bgcolor='rgba(35, 41, 70, 0.8)',
            bordercolor=colors['primary'],
            borderwidth=1
        ),
        font=dict(color=colors['text']),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(t=80, b=60, l=60, r=40),
        hovermode='x unified',
        autosize=True
    )
    
    # 多重强制设置Y轴
    fig.update_yaxes(
        range=[0, 100],
        autorange=False, 
        fixedrange=True,
        constrain='domain',
        dtick=20,
        tick0=0
    )
    
    # 最后一道保险：直接修改layout
    fig.layout.yaxis.range = [0, 100]
    fig.layout.yaxis.autorange = False
    
    return fig

def create_mutual_communication_trend_chart(trend_df):
    """创建相互沟通率趋势图"""
    # 创建图表，强制设置Y轴范围
    fig = go.Figure()
    
    # 添加一个隐藏的空数据点来强制Y轴范围
    fig.add_trace(go.Scatter(
        x=[],
        y=[0, 10],  # 强制Y轴从0到10
        mode='markers',
        marker=dict(opacity=0),  # 完全透明
        showlegend=False,
        hoverinfo='skip'
    ))
    
    if trend_df.empty:
        # 暂无数据时显示提示
        fig.add_annotation(
            text="暂无数据",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color=colors['text'])
        )
    else:
        # 添加相互沟通率线条
        fig.add_trace(go.Scatter(
            x=trend_df['date'],
            y=trend_df['mutual_communication_rate'],
            mode='lines+markers',
            name='相互沟通率',
            line=dict(color=colors['warning'], width=3),
            marker=dict(size=8, line=dict(width=2, color='white')),
            hovertemplate='<b>相互沟通率</b><br>日期: %{x}<br>沟通率: %{y:.1f}%<br>打招呼: %{customdata[0]}<br>相互沟通: %{customdata[1]}<extra></extra>',
            customdata=list(zip(trend_df['greetings'], trend_df['mutual_communications']))
        ))
    
    # 强制布局设置
    fig.update_layout(
        title=dict(
            text="<b>相互沟通率趋势分析</b>",
            font=dict(size=20, color=colors['text']),
            x=0.5
        ),
        xaxis=dict(
            title=dict(text="日期", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)'
        ),
        yaxis=dict(
            title=dict(text="沟通率 (%)", font=dict(color=colors['text'])),
            tickfont=dict(color=colors['text']),
            gridcolor='rgba(255,255,255,0.1)',
            range=[0, 10],
            fixedrange=True,
            autorange=False,
            tickformat='.1f',
            dtick=2,  # 每2%一个刻度
            tick0=0,  # 从0开始
            constrain='domain'
        ),
        legend=dict(
            font=dict(color=colors['text']),
            bgcolor='rgba(35, 41, 70, 0.8)',
            bordercolor=colors['primary'],
            borderwidth=1
        ),
        font=dict(color=colors['text']),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(t=80, b=60, l=60, r=40),
        hovermode='x unified',
        autosize=True
    )
    
    # 多重强制设置Y轴
    fig.update_yaxes(
        range=[0, 10],
        autorange=False, 
        fixedrange=True,
        constrain='domain',
        dtick=2,
        tick0=0
    )
    
    # 最后一道保险：直接修改layout
    fig.layout.yaxis.range = [0, 10]
    fig.layout.yaxis.autorange = False
    
    return fig

# 应用布局
app.layout = html.Div([
    # 隐藏的存储组件用于状态管理
    dcc.Store(id='last-refresh-time'),
    dcc.Store(id='current-data'),
    
    # 头部
    html.Div([
        html.Div([
            html.Div([
                html.H1("🚀 智能招聘数据分析平台"),
                html.P("实时监控 · 数据驱动 · 智能决策")
            ], className="logo-section"),
            
            html.Div([
                html.Span(id="last-update-time", children="数据加载中..."),
                html.Button(
                    "🔄 刷新数据", 
                    id="refresh-btn",
                    className="btn",
                    style={'marginLeft': '15px'}
                )
            ], style={'display': 'flex', 'alignItems': 'center'})
        ], className="header-content")
    ], className="dashboard-header"),
    
    # 控制面板
    html.Div([
        html.Div([
            html.Div([
                html.Label("📅 选择日期范围"),
                dcc.DatePickerRange(
                    id='date-picker-range',
                    start_date=date.today() - timedelta(days=30),
                    end_date=date.today(),
                    display_format='YYYY-MM-DD',
                    style={'width': '100%'},
                    start_date_placeholder_text="开始日期",
                    end_date_placeholder_text="结束日期",
                    clearable=True
                )
            ], className="control-group"),
            
            html.Div([
                html.Label("👤 选择用户"),
                dcc.Dropdown(
                    id='user-dropdown',
                    options=get_user_list(),
                    value='all',
                    clearable=False,
                    style={'minWidth': '200px'}
                )
            ], className="control-group"),
            
            html.Div([
                html.Label("🔄 自动刷新"),
                dcc.Dropdown(
                    id='refresh-interval',
                    options=[
                        {'label': '关闭', 'value': 0},
                        {'label': '30秒', 'value': 30000},
                        {'label': '1分钟', 'value': 60000},
                        {'label': '5分钟', 'value': 300000}
                    ],
                    value=0,
                    clearable=False
                )
            ], className="control-group")
        ], className="control-content")
    ], className="control-panel"),
    
    # 主要内容
    html.Div([
        # KPI指标卡片区域
        html.Div(id="kpi-section", className="kpi-section"),
        
        # 图表区域
        html.Div([
            html.Div([
                html.Div([
                    html.Div([
                        html.H2("📊 招聘漏斗分析", className="card-title"),
                        html.Div(id="funnel-loading", children=create_loading_component())
                    ], className="card-header"),
                    dcc.Graph(id="funnel-chart", style={'display': 'none'})
                ], className="card")
            ], style={'gridColumn': '1'}),
            
            html.Div([
                html.Div([
                    html.Div([
                        html.H2("📈 每日活动趋势", className="card-title"),
                        html.Div(id="trend-loading", children=create_loading_component())
                    ], className="card-header"),
                    dcc.Graph(id="trend-chart", style={'display': 'none'})
                ], className="card")
            ], style={'gridColumn': '2'})
        ], className="content-grid"),
        
        # 新增打招呼成功率趋势图
        html.Div([
            html.Div([
                html.Div([
                    html.H2("📈 打招呼成功率趋势分析", className="card-title"),
                    html.Div(id="success-trend-loading", children=create_loading_component())
                ], className="card-header"),
                dcc.Graph(id="success-trend-chart", style={'display': 'none'})
            ], className="card")
        ], style={'marginBottom': '2.5rem'}),
        
        # 新增沟通成功率和相互沟通率趋势图
        html.Div([
            html.Div([
                html.Div([
                    html.Div([
                        html.H2("🎯 沟通成功率趋势分析", className="card-title"),
                        html.Div(id="comm-success-trend-loading", children=create_loading_component())
                    ], className="card-header"),
                    dcc.Graph(id="comm-success-trend-chart", style={'display': 'none'})
                ], className="card")
            ], style={'gridColumn': '1'}),
            
            html.Div([
                html.Div([
                    html.Div([
                        html.H2("💡 相互沟通率趋势分析", className="card-title"),
                        html.Div(id="mutual-comm-trend-loading", children=create_loading_component())
                    ], className="card-header"),
                    dcc.Graph(id="mutual-comm-trend-chart", style={'display': 'none'})
                ], className="card")
            ], style={'gridColumn': '2'})
        ], className="content-grid", style={'marginBottom': '2.5rem'}),
        
        # 详细数据表格
        html.Div([
            html.Div([
                html.Div([
                    html.H2("📋 详细数据", className="card-title"),
                    html.Div([
                        html.Button("📊 导出Excel", id="export-excel-btn", className="btn btn-outline", style={'marginRight': '10px'}),
                        html.Button("📄 导出CSV", id="export-csv-btn", className="btn btn-outline")
                    ])
                ], className="card-header"),
                html.Div(id="data-table-loading", children=create_loading_component()),
                html.Div(id="data-table-container", style={'display': 'none'})
            ], className="card")
        ], className="data-table-container")
    ], className="main-content"),
    
    # 自动刷新组件
    dcc.Interval(
        id='auto-refresh-interval',
        interval=60000,  # 默认1分钟
        n_intervals=0,
        disabled=True
    ),
    
    # 下载组件
    dcc.Download(id="download-excel"),
    dcc.Download(id="download-csv")
], className="dashboard-container")

# 快速筛选回调已删除

# 回调函数：自动刷新设置
@app.callback(
    [Output('auto-refresh-interval', 'interval'),
     Output('auto-refresh-interval', 'disabled')],
    [Input('refresh-interval', 'value')]
)
def update_auto_refresh(interval_value):
    if interval_value == 0:
        return 60000, True
    else:
        return interval_value, False

# 回调函数：更新数据
@app.callback(
    [Output('kpi-section', 'children'),
     Output('funnel-chart', 'figure'),
     Output('funnel-chart', 'style'),
     Output('funnel-loading', 'style'),
     Output('trend-chart', 'figure'),
     Output('trend-chart', 'style'),
     Output('trend-loading', 'style'),
     Output('success-trend-chart', 'figure'),
     Output('success-trend-chart', 'style'),
     Output('success-trend-loading', 'style'),
     Output('comm-success-trend-chart', 'figure'),
     Output('comm-success-trend-chart', 'style'),
     Output('comm-success-trend-loading', 'style'),
     Output('mutual-comm-trend-chart', 'figure'),
     Output('mutual-comm-trend-chart', 'style'),
     Output('mutual-comm-trend-loading', 'style'),
     Output('data-table-container', 'children'),
     Output('data-table-container', 'style'),
     Output('data-table-loading', 'style'),
     Output('last-update-time', 'children'),
     Output('last-refresh-time', 'data')],
    [Input('date-picker-range', 'start_date'),
     Input('date-picker-range', 'end_date'),
     Input('user-dropdown', 'value'),
     Input('refresh-btn', 'n_clicks'),
     Input('auto-refresh-interval', 'n_intervals')]
)
def update_dashboard(start_date, end_date, user_id, refresh_clicks, auto_refresh):
    # 显示加载状态
    loading_style = {'display': 'block'}
    hidden_style = {'display': 'none'}
    
    try:
        # 获取数据
        metrics = get_key_metrics(start_date, end_date, user_id)
        funnel_df = get_funnel_data(start_date, end_date, user_id)
        trend_df = get_trend_data(start_date, end_date, user_id)
        success_trend_df = get_greeting_success_trend(start_date, end_date, user_id)
        comm_success_trend_df = get_communication_success_trend(start_date, end_date, user_id)
        mutual_comm_trend_df = get_mutual_communication_trend(start_date, end_date, user_id)
        detailed_df = get_detailed_data(start_date, end_date, user_id, limit=1000)
        
        # 创建KPI卡片（包含计算方式）
        kpi_cards = [
            create_metric_card("📊 浏览简历", metrics['browse_resumes'], None, 'number'),
            create_metric_card("✅ 打招呼", metrics['greetings'], None, 'number'),
            create_metric_card("💬 相互沟通", metrics['mutual_communications'], None, 'number'),
            create_metric_card("🤝 建联量", metrics['connections'], None, 'number'),
            create_metric_card("📈 打招呼成功率", metrics['greeting_success_rate'], None, 'percentage', 
                             None, "计算方式：建联量 ÷ 打招呼数 × 100%"),
            create_metric_card("🎯 沟通成功率", metrics['communication_success_rate'], None, 'percentage', 
                             None, "计算方式：建联量 ÷ 相互沟通数 × 100%"),
            create_metric_card("💡 相互沟通率", metrics['mutual_communication_rate'], None, 'percentage', 
                             None, "计算方式：相互沟通数 ÷ 打招呼数 × 100%"),
            create_metric_card("📋 简历过筛率", metrics['resume_screening_rate'], None, 'percentage', 
                             None, "计算方式：打招呼数 ÷ 浏览简历数 × 100%")
        ]
        
        # 创建图表
        funnel_chart = create_funnel_chart(funnel_df)
        trend_chart = create_trend_chart(trend_df)
        success_trend_chart = create_greeting_success_trend_chart(success_trend_df)
        comm_success_trend_chart = create_communication_success_trend_chart(comm_success_trend_df)
        mutual_comm_trend_chart = create_mutual_communication_trend_chart(mutual_comm_trend_df)
        
        # 创建数据表格
        if not detailed_df.empty:
            # 移除ID列用于显示
            display_df = detailed_df.drop(columns=['id'] if 'id' in detailed_df.columns else [])
            
            data_table = dash_table.DataTable(
                data=display_df.to_dict('records'),
                columns=[{"name": i, "id": i} for i in display_df.columns],
                page_size=20,
                sort_action="native",
                filter_action="native",
                style_table={'overflowX': 'auto'},
                style_cell={
                    'textAlign': 'left',
                    'padding': '12px',
                    'fontFamily': 'Arial',
                    'fontSize': '14px'
                },
                style_header={
                    'fontWeight': 'bold',
                    'textAlign': 'center'
                },
                style_data_conditional=[
                    {
                        'if': {'row_index': 'odd'},
                        'backgroundColor': 'rgba(30, 39, 73, 0.5)'
                    }
                ]
            )
        else:
            data_table = html.Div("暂无数据", style={
                'textAlign': 'center', 
                'padding': '40px',
                'color': colors['text'],
                'fontSize': '16px'
            })
        
        # 更新时间戳
        update_time = f"最后更新: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        return (
            kpi_cards,
            funnel_chart, {'display': 'block'}, hidden_style,
            trend_chart, {'display': 'block'}, hidden_style,
            success_trend_chart, {'display': 'block'}, hidden_style,
            comm_success_trend_chart, {'display': 'block'}, hidden_style,
            mutual_comm_trend_chart, {'display': 'block'}, hidden_style,
            data_table, {'display': 'block'}, hidden_style,
            update_time,
            time.time()
        )
        
    except Exception as e:
        print(f"更新数据时出错: {e}")
        error_msg = html.Div(f"数据加载失败: {str(e)}", style={
            'color': colors['danger'],
            'textAlign': 'center',
            'padding': '20px'
        })
        
        return (
            [error_msg],
            {}, hidden_style, hidden_style,
            {}, hidden_style, hidden_style,
            {}, hidden_style, hidden_style,
            {}, hidden_style, hidden_style,
            {}, hidden_style, hidden_style,
            error_msg, hidden_style, hidden_style,
            f"更新失败: {datetime.now().strftime('%H:%M:%S')}",
            time.time()
        )

# 导出Excel回调
@app.callback(
    Output('download-excel', 'data'),
    [Input('export-excel-btn', 'n_clicks')],
    [State('date-picker-range', 'start_date'),
     State('date-picker-range', 'end_date'),
     State('user-dropdown', 'value')],
    prevent_initial_call=True
)
def export_excel(n_clicks, start_date, end_date, user_id):
    if n_clicks:
        try:
            # 获取所有数据（不限制条数）
            raw_data_df = get_detailed_data(start_date, end_date, user_id, limit=None)
            metrics = get_key_metrics(start_date, end_date, user_id)
            success_trend_df = get_greeting_success_trend(start_date, end_date, user_id)
            comm_success_trend_df = get_communication_success_trend(start_date, end_date, user_id)
            mutual_comm_trend_df = get_mutual_communication_trend(start_date, end_date, user_id)
            funnel_df = get_funnel_data(start_date, end_date, user_id)
            trend_df = get_trend_data(start_date, end_date, user_id)
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"智能招聘数据分析_{timestamp}.xlsx"
            
            # 创建Excel文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # 1. 原始数据表
                if not raw_data_df.empty:
                    display_df = raw_data_df.drop(columns=['id'] if 'id' in raw_data_df.columns else [])
                    display_df.to_excel(writer, sheet_name='01_原始数据', index=False)
                
                # 2. KPI指标汇总表（使用真实的Excel公式）
                kpi_data = {
                    'KPI指标': [
                        '浏览简历数',
                        '打招呼数', 
                        '相互沟通数',
                        '建联量',
                        '打招呼成功率(%)',
                        '沟通成功率(%)',
                        '相互沟通率(%)',
                        '简历过筛率(%)'
                    ],
                    '数值': [
                        metrics['browse_resumes'],
                        metrics['greetings'],
                        metrics['mutual_communications'],
                        metrics['connections'],
                        '=IF(B3=0,0,B5/B3*100)',  # 使用真实的Excel公式
                        '=IF(B4=0,0,B5/B4*100)', 
                        '=IF(B3=0,0,B4/B3*100)',
                        '=IF(B2=0,0,B3/B2*100)'
                    ],
                    '计算公式说明': [
                        '直接统计',
                        '直接统计',
                        '直接统计', 
                        '直接统计',
                        '建联量÷打招呼数×100%',
                        '建联量÷相互沟通数×100%',
                        '相互沟通数÷打招呼数×100%',
                        '打招呼数÷浏览简历数×100%'
                    ],
                    '业务含义': [
                        '事件类型=1的记录数',
                        '事件类型=2的记录数',
                        '事件类型=12的记录数',
                        '事件类型=13的记录数',
                        '衡量打招呼后建立连接的效率',
                        '衡量沟通后建立连接的效率',
                        '衡量打招呼后获得回应的比例',
                        '衡量简历筛选后主动联系的比例'
                    ]
                }
                kpi_df = pd.DataFrame(kpi_data)
                kpi_df.to_excel(writer, sheet_name='02_KPI指标汇总', index=False)
                
                # 3. 漏斗分析数据
                if not funnel_df.empty:
                    funnel_data = funnel_df.copy()
                    # 添加转化率计算公式
                    funnel_data['总体转化率(%)'] = [
                        '=B2/B2*100',  # 浏览简历基准
                        '=IF(B2=0,0,B3/B2*100)',  # 打招呼/浏览简历
                        '=IF(B2=0,0,B4/B2*100)',  # 相互沟通/浏览简历
                        '=IF(B2=0,0,B5/B2*100)'   # 建联量/浏览简历
                    ]
                    funnel_data['上级转化率(%)'] = [
                        '=B2/B2*100',  # 浏览简历基准
                        '=IF(B2=0,0,B3/B2*100)',  # 打招呼/浏览简历
                        '=IF(B3=0,0,B4/B3*100)',  # 相互沟通/打招呼
                        '=IF(B4=0,0,B5/B4*100)'   # 建联量/相互沟通
                    ]
                    funnel_data.to_excel(writer, sheet_name='03_漏斗分析数据', index=False)
                    
                # 4. 每日活动趋势
                if not trend_df.empty:
                    trend_pivot = trend_df.pivot_table(
                        index='date', 
                        columns='event_type', 
                        values='count', 
                        fill_value=0
                    ).reset_index()
                    trend_pivot.to_excel(writer, sheet_name='04_每日活动趋势', index=False)
                
                # 5. 打招呼成功率趋势
                if not success_trend_df.empty:
                    trend_data = success_trend_df.copy()
                    # 使用真实的Excel公式
                    trend_data['成功率公式验证'] = [f'=IF(C{i+2}=0,0,D{i+2}/C{i+2}*100)' for i in range(len(trend_data))]
                    trend_data.to_excel(writer, sheet_name='05_打招呼成功率趋势', index=False)
                
                # 6. 沟通成功率趋势
                if not comm_success_trend_df.empty:
                    comm_trend_data = comm_success_trend_df.copy()
                    # 使用真实的Excel公式
                    comm_trend_data['成功率公式验证'] = [f'=IF(C{i+2}=0,0,D{i+2}/C{i+2}*100)' for i in range(len(comm_trend_data))]  
                    comm_trend_data.to_excel(writer, sheet_name='06_沟通成功率趋势', index=False)
                
                # 7. 相互沟通率趋势
                if not mutual_comm_trend_df.empty:
                    mutual_trend_data = mutual_comm_trend_df.copy()
                    # 使用真实的Excel公式
                    mutual_trend_data['沟通率公式验证'] = [f'=IF(C{i+2}=0,0,D{i+2}/C{i+2}*100)' for i in range(len(mutual_trend_data))]
                    mutual_trend_data.to_excel(writer, sheet_name='07_相互沟通率趋势', index=False)
                
                # 8. 图表数据汇总（用于创建图表）
                if not trend_df.empty and not success_trend_df.empty:
                    # 创建综合的图表数据表
                    chart_summary = []
                    
                    # 基础活动数据（来自网页折线图）
                    trend_pivot = trend_df.pivot_table(
                        index='date', 
                        columns='event_type', 
                        values='count', 
                        fill_value=0
                    ).reset_index()
                    
                    # 合并所有趋势数据
                    chart_data = trend_pivot.copy()
                    if not success_trend_df.empty:
                        chart_data = chart_data.merge(success_trend_df[['date', 'greeting_success_rate']], on='date', how='outer')
                    if not comm_success_trend_df.empty:
                        chart_data = chart_data.merge(comm_success_trend_df[['date', 'communication_success_rate']], on='date', how='outer')
                    if not mutual_comm_trend_df.empty:
                        chart_data = chart_data.merge(mutual_comm_trend_df[['date', 'mutual_communication_rate']], on='date', how='outer')
                    
                    # 填充空值
                    chart_data = chart_data.fillna(0)
                    chart_data = chart_data.sort_values('date')
                    
                    # 重新排列列的顺序，使其更直观
                    cols = ['date']
                    event_cols = [col for col in chart_data.columns if col not in ['date', 'greeting_success_rate', 'communication_success_rate', 'mutual_communication_rate']]
                    rate_cols = ['greeting_success_rate', 'communication_success_rate', 'mutual_communication_rate']
                    chart_data = chart_data[cols + event_cols + rate_cols]
                    
                    chart_data.to_excel(writer, sheet_name='08_图表数据汇总', index=False)
                
                # 9. 数据字典说明
                dict_data = {
                    '字段名': [
                        'event_type',
                        'event_type=1',
                        'event_type=2', 
                        'event_type=12',
                        'event_type=13',
                        'create_time',
                        'uid',
                        'user_name',
                        'resume_id',
                        'job_id'
                    ],
                    '含义': [
                        '事件类型',
                        '浏览简历',
                        '打招呼',
                        '相互沟通',
                        '建联量',
                        '事件发生时间',
                        '用户ID',
                        '用户姓名',
                        '简历ID',
                        '职位ID'
                    ],
                    '说明': [
                        '区分不同类型的招聘行为',
                        'HR查看候选人简历',
                        'HR主动向候选人打招呼',
                        'HR与候选人进行相互交流',
                        'HR成功与候选人建立联系',
                        '精确到秒的时间戳',
                        '用户唯一标识符',
                        '用户真实姓名或生成的显示名',
                        '被查看简历的唯一标识',
                        '相关职位的唯一标识'
                    ]
                }
                dict_df = pd.DataFrame(dict_data)
                dict_df.to_excel(writer, sheet_name='09_数据字典', index=False)
                
                # 10. 计算公式说明
                formula_data = {
                    '指标名称': [
                        '打招呼成功率',
                        '沟通成功率',
                        '相互沟通率',
                        '简历过筛率',
                        '总体转化率',
                        '上级转化率'
                    ],
                    'Excel公式模板': [
                        '=IF(打招呼数=0,0,建联量/打招呼数*100)',
                        '=IF(相互沟通数=0,0,建联量/相互沟通数*100)',
                        '=IF(打招呼数=0,0,相互沟通数/打招呼数*100)',
                        '=IF(浏览简历数=0,0,打招呼数/浏览简历数*100)',
                        '以浏览简历为基准计算各阶段占比',
                        '以上一级为基准计算转化率'
                    ],
                    '业务含义': [
                        '衡量打招呼后建立连接的效率',
                        '衡量沟通后建立连接的效率', 
                        '衡量打招呼后获得回应的比例',
                        '衡量简历筛选后主动联系的比例',
                        '衡量整体流程中各环节的绝对效率',
                        '衡量相邻环节间的转化效率'
                    ],
                    '优化建议': [
                        '提高个人资料完整度，优化打招呼话术',
                        '改进沟通技巧，提供有价值的信息',
                        '优化打招呼时机和内容吸引力',
                        '提高简历筛选精准度，避免无效联系',
                        '关注整体流程优化，减少各环节流失',
                        '重点关注转化率低的相邻环节'
                    ]
                }
                formula_df = pd.DataFrame(formula_data)
                formula_df.to_excel(writer, sheet_name='10_公式说明', index=False)
                
                # 11. 使用说明
                usage_data = {
                    '表格名称': [
                        '01_原始数据',
                        '02_KPI指标汇总',
                        '03_漏斗分析数据',
                        '04_每日活动趋势',
                        '05_打招呼成功率趋势',
                        '06_沟通成功率趋势',
                        '07_相互沟通率趋势',
                        '08_图表数据汇总',
                        '09_数据字典',
                        '10_公式说明'
                    ],
                    '用途说明': [
                        '完整的原始数据，可用于进一步分析',
                        '关键指标汇总，B列包含Excel公式自动计算',
                        '漏斗分析，包含转化率的Excel公式计算',
                        '每日各类活动的统计数据',
                        '每日打招呼成功率变化趋势',
                        '每日沟通成功率变化趋势',
                        '每日相互沟通率变化趋势',
                        '所有图表数据的汇总，可用于制作Excel图表',
                        '数据字段的详细解释',
                        '所有计算公式的说明和优化建议'
                    ],
                    'Excel功能': [
                        '支持筛选、排序、数据透视表',
                        '公式自动计算，数据更新时同步更新',
                        '转化率公式自动计算',
                        '可制作柱状图、折线图',
                        '可制作趋势折线图',
                        '可制作趋势折线图',
                        '可制作趋势折线图',
                        '推荐制作综合仪表板图表',
                        '参考查询，不建议修改',
                        '参考查询，包含优化建议'
                    ],
                    '建议操作': [
                        '按日期、用户、事件类型筛选分析',
                        '监控各项指标变化，关注公式计算结果',
                        '分析转化率，识别优化机会',
                        '制作活动趋势图表，观察周期性规律',
                        '制作成功率趋势图，找出最佳时机',
                        '制作成功率趋势图，优化沟通策略',
                        '制作沟通率趋势图，提高回应率',
                        '制作综合数据看板，一次性展示所有指标',
                        '了解数据含义，正确解读分析结果',
                        '参考优化建议，制定改进措施'
                    ]
                }
                usage_df = pd.DataFrame(usage_data)
                usage_df.to_excel(writer, sheet_name='00_使用说明', index=False)
                
                # 设置Excel样式
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                workbook = writer.book
                
                # 定义样式
                header_font = Font(bold=True, size=12, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                center_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'), 
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 为每个工作表设置样式
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # 设置表头样式
                    if worksheet.max_row > 0:
                        for cell in worksheet[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = thin_border
                    
                    # 自动调整列宽
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
            output.seek(0)
            
            # 触发下载
            return dcc.send_bytes(output.getvalue(), filename)
            
        except Exception as e:
            print(f"导出Excel失败: {e}")
            import traceback
            traceback.print_exc()
    
    return dash.no_update

# 导出CSV回调
@app.callback(
    Output('download-csv', 'data'),
    [Input('export-csv-btn', 'n_clicks')],
    [State('date-picker-range', 'start_date'),
     State('date-picker-range', 'end_date'),
     State('user-dropdown', 'value')],
    prevent_initial_call=True
)
def export_csv(n_clicks, start_date, end_date, user_id):
    if n_clicks:
        try:
            # 获取详细数据
            df = get_detailed_data(start_date, end_date, user_id)
            if not df.empty:
                # 移除ID列
                if 'id' in df.columns:
                    df = df.drop(columns=['id'])
                
                # 生成文件名
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"招聘数据_{timestamp}.csv"
                
                # 生成CSV
                csv_string = df.to_csv(index=False, encoding='utf-8-sig')
                
                # 触发下载
                return dcc.send_string(csv_string, filename)
        except Exception as e:
            print(f"导出CSV失败: {e}")
    
    return dash.no_update

if __name__ == '__main__':
    print("🚀 启动智能招聘数据分析平台...")
    print("📊 访问地址: http://0.0.0.0:8050/")
    print("🌐 局域网访问: http://172.18.202.183:8050/")
    app.run(host='0.0.0.0', port=8050, debug=True) 